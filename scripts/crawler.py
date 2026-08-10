# -*- coding: utf-8 -*-
"""
爬虫展示案例：中国政府采购网 · 中央公告采集
=============================================
自动抓取"中央公告"列表页，解析出 标题/公告类型/发布时间/地域/采购人/链接，
去重合并后输出 CSV + Excel 报表 + 采集摘要。

仅采集公开信息，低频访问，不涉及任何个人信息。
用法（在 frontend-showcase 根目录运行）:
    python scripts/crawler.py                     # 用 scripts/config.json 默认配置
    python scripts/crawler.py --pages 2           # 抓 2 页
    python scripts/crawler.py --delay 1.5         # 覆盖访问间隔（秒）
    python scripts/crawler.py --outdir output     # 覆盖输出目录
    python scripts/crawler.py --export-events crawler-showcase/public/events.json
                                                  # 额外导出过程事件流（网页重演数据源）
    python scripts/crawler.py --limit 5           # 最多采集 5 条（数据量自定义）
"""

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:  # Excel 输出为可选，缺失时仅保留 CSV
    openpyxl = None

DEFAULT_CONFIG = {
    "source": {
        "name": "中国政府采购网-中央公告",
        "list_url": "https://www.ccgp.gov.cn/cggg/zygg/",
        "encoding": "utf-8",
        "pages": 1,
        "delay_seconds": 1.0,
        "headers": {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        },
        "retries": 2,
        "timeout": 20,
    },
    "output": {
        "dir": "output",
        "csv": "政采公告_中央.csv",
        "xlsx": "政采公告_中央.xlsx",
        "summary": "采集摘要.txt",
    },
}


def load_config(path: Path) -> dict:
    cfg = json.loads(path.read_text(encoding="utf-8"))
    # 与默认配置合并，保证字段齐全
    merged = json.loads(json.dumps(DEFAULT_CONFIG))
    for sec in cfg:
        if isinstance(cfg[sec], dict):
            merged[sec].update(cfg[sec])
        else:
            merged[sec] = cfg[sec]
    return merged


SCRIPT_DIR = Path(__file__).resolve().parent   # scripts/
PROJECT_DIR = SCRIPT_DIR.parent                 # frontend-showcase/


class EventRecorder:
    """过程事件流记录器：把爬取过程按时间顺序记成 JSON，供网页重演播放。"""

    def __init__(self, path: Path):
        self.path = path
        self.events = []
        self.t0 = time.time()

    def log(self, etype: str, msg: str, **data):
        ev = {"t": round(time.time() - self.t0, 3), "type": etype, "msg": msg}
        if data:
            ev["data"] = data
        self.events.append(ev)

    def save(self, meta: dict):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        payload = {"meta": meta, "events": self.events}
        self.path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch(session: requests.Session, url: str, cfg: dict, rec: EventRecorder = None) -> str:
    """带重试的抓取，返回解码后的 HTML 文本。"""
    last_err = None
    for attempt in range(cfg["source"]["retries"] + 1):
        try:
            r = session.get(url, timeout=cfg["source"]["timeout"])
            if r.status_code == 404:
                if rec:
                    rec.log("request", f"GET {url} → 404（翻页到底）")
                return None  # 翻页到底
            r.raise_for_status()
            r.encoding = cfg["source"].get("encoding", "utf-8")
            if rec:
                rec.log("request", f"GET {url} → 200（{len(r.text)} bytes）")
            return r.text
        except requests.RequestException as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    if rec:
        rec.log("request_failed", f"GET {url} 失败：{last_err}")
    print(f"  [warn] fetch failed: {url} ({last_err})", flush=True)
    return None


def parse_items(html: str) -> list:
    """解析列表页，返回 [{title, type, pub_time, region, buyer, url}]"""
    soup = BeautifulSoup(html, "html.parser")
    items = []
    for li in soup.select("ul.c_list_bid > li"):
        a = li.find("a", href=True)
        if not a:
            continue
        title = (a.get("title") or a.get_text(strip=True) or "").strip()
        if not title:
            continue
        btype = li.select_one('em[rel="bxlx"]')
        btype = btype.get_text(strip=True) if btype else ""
        text = re.sub(r"\s+", " ", li.get_text(" ", strip=True))
        m = re.search(r"发布时间：(.+?) 地域：(.+?) 采购人：(.+)$", text)
        pub_time = m.group(1).strip() if m else ""
        region = m.group(2).strip() if m else ""
        buyer = m.group(3).strip() if m else ""
        items.append({
            "title": title,
            "type": btype,
            "pub_time": pub_time,
            "region": region,
            "buyer": buyer,
            "url": a["href"],
        })
    return items


def to_absolute(base: str, href: str) -> str:
    from urllib.parse import urljoin
    return urljoin(base if base.endswith("/") else base + "/", href)


def page_url(base: str, page: int) -> str:
    """第 1 页用列表页本身；第 N 页用 index_N.htm（ccgp 翻页约定）。"""
    if page <= 1:
        return base
    return base.rstrip("/") + f"/index_{page}.htm"


def load_existing_links(csv_path: Path) -> set:
    if not csv_path.exists():
        return set()
    try:
        import csv as _csv
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            return {row.get("链接", "") for row in _csv.DictReader(f)}
    except Exception:
        return set()


def write_csv(path: Path, rows: list):
    import csv as _csv
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = _csv.DictWriter(f, fieldnames=["标题", "公告类型", "发布时间", "地域", "采购人", "链接"])
        w.writeheader()
        for r in rows:
            w.writerow({
                "标题": r["title"],
                "公告类型": r["type"],
                "发布时间": r["pub_time"],
                "地域": r["region"],
                "采购人": r["buyer"],
                "链接": r["url"],
            })


def write_xlsx(path: Path, rows: list):
    if openpyxl is None:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "政采公告"
    headers = ["标题", "公告类型", "发布时间", "地域", "采购人", "链接"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([r["title"], r["type"], r["pub_time"], r["region"], r["buyer"], r["url"]])
    widths = [70, 12, 18, 8, 40, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"
    wb.save(path)


def write_summary(path: Path, rows: list, fetched: int, new_count: int, elapsed: float):
    from collections import Counter
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"采集时间：{now}",
        f"数据源：{rows[0]['url'].split('/')[2] if rows else ''} 中央公告列表",
        f"本次抓取 {fetched} 条，新增 {new_count} 条，累计 {len(rows)} 条（按链接去重）",
        f"耗时 {elapsed:.1f} 秒",
        "",
        "按公告类型分布：",
    ]
    for k, v in Counter(r["type"] or "未知" for r in rows).most_common():
        lines.append(f"  {k}: {v}")
    lines.append("")
    lines.append("按地域分布（前 10）：")
    for k, v in Counter(r["region"] or "未知" for r in rows).most_common(10):
        lines.append(f"  {k}: {v}")
    lines.append("")
    lines.append("最近 5 条：")
    for r in rows[:5]:
        lines.append(f"  [{r['pub_time']}] {r['title']}（{r['region']} · {r['buyer']}）")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    ap = argparse.ArgumentParser(description="政采公告采集展示案例")
    ap.add_argument("--config", default=None, help="配置文件路径（默认 scripts/config.json）")
    ap.add_argument("--pages", type=int, default=None, help="抓取页数（覆盖配置）")
    ap.add_argument("--delay", type=float, default=None, help="访问间隔秒数（覆盖配置）")
    ap.add_argument("--outdir", default=None, help="输出目录（覆盖配置）")
    ap.add_argument("--export-events", default=None,
                    help="导出过程事件流 JSON（网页重演数据源）")
    ap.add_argument("--limit", type=int, default=None, help="最多采集条数（数据量自定义）")
    args = ap.parse_args()

    cfg = load_config(Path(args.config) if args.config else SCRIPT_DIR / "config.json")
    src, out = cfg["source"], cfg["output"]
    pages = args.pages if args.pages is not None else src["pages"]
    delay = args.delay if args.delay is not None else src["delay_seconds"]
    outdir = Path(args.outdir) if args.outdir else PROJECT_DIR / out["dir"]
    outdir.mkdir(parents=True, exist_ok=True)
    csv_path = outdir / out["csv"]
    xlsx_path = outdir / out["xlsx"]
    summary_path = outdir / out["summary"]
    rec = EventRecorder(Path(args.export_events)) if args.export_events else None

    session = requests.Session()
    session.headers.update(src["headers"])

    existing_links = load_existing_links(csv_path)
    print(f"[1/3] fetch list pages x{pages}: {src['list_url']}", flush=True)
    t0 = time.time()
    new_items = []
    seen = set()
    fetched_total = 0
    collected = 0
    for page in range(1, pages + 1):
        url = page_url(src["list_url"], page)
        html = fetch(session, url, cfg, rec)
        if not html:
            if page > 1:
                print(f"  [info] no more pages at {url}", flush=True)
            break
        items = parse_items(html)
        if args.limit is not None:
            remain = args.limit - collected
            if remain <= 0:
                break
            items = items[:remain]
        collected += len(items)
        fetched_total += len(items)
        if rec:
            rec.log("parse", f"解析 {url}：{len(items)} 条公告")
        for n, it in enumerate(items, 1):
            it["url"] = to_absolute(url, it["url"])
            if rec:
                rec.log("item", f"[{it['type'] or '公告'}] {it['title']}", item=it)
        fresh = [it for it in items if it["url"] not in existing_links and it["url"] not in seen]
        for it in fresh:
            seen.add(it["url"])
        new_items.extend(fresh)
        print(f"  page {page}: {len(items)} items, {len(fresh)} new", flush=True)
        if page < pages:
            time.sleep(delay)

    # 合并历史数据（按链接去重，新条目在前）
    rows = []
    if csv_path.exists():
        import csv as _csv
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            for row in _csv.DictReader(f):
                rows.append({
                    "title": row["标题"], "type": row["公告类型"],
                    "pub_time": row["发布时间"], "region": row["地域"],
                    "buyer": row["采购人"], "url": row["链接"],
                })
    rows = new_items + [r for r in rows if r["url"] not in seen]

    if rec:
        rec.log("dedupe", f"按链接去重：抓取 {fetched_total} 条，其中新增 {len(new_items)} 条")

    write_csv(csv_path, rows)
    write_xlsx(xlsx_path, rows)
    elapsed = time.time() - t0
    write_summary(summary_path, rows, fetched_total, len(new_items), elapsed)

    if rec:
        rec.log("export", f"生成报表：{out['xlsx']} / {out['csv']}（共 {len(rows)} 行）")
        rec.log("done", f"完成：累计 {len(rows)} 条，耗时 {elapsed:.1f}s")
        rec.save({
            "source": src["name"],
            "list_url": src["list_url"],
            "run_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fetched": fetched_total,
            "new": len(new_items),
            "total": len(rows),
            "elapsed": round(elapsed, 1),
        })
        print(f"events: {rec.path}", flush=True)

    print(f"[2/3] saved CSV: {csv_path}", flush=True)
    if xlsx_path.exists():
        print(f"[3/3] saved XLSX: {xlsx_path}", flush=True)
    print(f"done: total {len(rows)} rows (new {len(new_items)}), {elapsed:.1f}s", flush=True)
    print(f"summary: {summary_path}", flush=True)


if __name__ == "__main__":
    main()
